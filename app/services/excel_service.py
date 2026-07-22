import os
import re
from copy import copy
from io import BytesIO
from itertools import islice
from openpyxl import load_workbook


def _normalize_header(header):
    if header is None:
        return ""
    text = str(header).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def _resolve_field_for_header(header):
    """Maps a column header to the primary record field it represents, or
    None if it should be treated as a free-form extra_data column."""
    normalized = _normalize_header(header)

    is_pid_remarks = (
        normalized in {"remarks_in_pid", "remarks_in_pid_database", "remarks_pid", "database_remarks"}
        or (
            ("remark" in normalized or "remak" in normalized)
            and ("pid" in normalized or "p_id" in normalized or "database" in normalized)
        )
    )

    if is_pid_remarks:
        return "remarks_in_pid"
    elif normalized in {"boccard_item_number", "boccard", "boccard_item", "item_number"} or "boccard" in normalized:
        return "boccard_item_number"
    elif normalized in {"tag_number", "tag", "tag_no", "tagnumber"} or "tag" in normalized:
        return "tag_number"
    elif normalized in {"designation", "design", "desig"} or "designation" in normalized:
        return "designation"
    return None


def _map_value(record, header, value):
    if value is None:
        return

    text = str(value).strip()
    if not text:
        return

    field = _resolve_field_for_header(header)
    if field:
        record[field] = text
    else:
        record["extra_data"][header] = text


def _score_header_row(row):
    score = 0
    for value in row:
        normalized = _normalize_header(value)
        if not normalized:
            continue

        if "remark" in normalized or "remak" in normalized or "pid" in normalized:
            score += 2
        if "boccard" in normalized or "item_number" in normalized:
            score += 2
        if normalized in {"tag", "tag_no", "tag_number", "tagnumber"} or "tag" in normalized:
            score += 2
        if "designation" in normalized:
            score += 2

    return score


def _find_header_index(rows):
    best_index = 0
    best_score = 0

    for index, row in enumerate(rows[:25]):
        score = _score_header_row(row)
        if score > best_score:
            best_index = index
            best_score = score

    return best_index


def _find_best_sheet(workbook):
    best_sheet = workbook.active
    best_header_index = 0
    best_score = 0

    for sheet in workbook.worksheets:
        sample_rows = list(islice(sheet.iter_rows(values_only=True), 25))
        if not sample_rows:
            continue

        header_index = _find_header_index(sample_rows)
        score = _score_header_row(sample_rows[header_index])
        if score > best_score:
            best_sheet = sheet
            best_header_index = header_index
            best_score = score

    rows = list(best_sheet.iter_rows(values_only=True))
    return best_sheet, rows, best_header_index


def parse_excel_file(path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    _sheet, rows, header_index = _find_best_sheet(workbook)
    if not rows:
        return []

    headers = [str(header).strip() if header is not None else "" for header in rows[header_index]]
    parsed_rows = []

    for row in rows[header_index + 1:]:
        if not row or all(value is None or str(value).strip() == "" for value in row):
            continue

        record = {
            "remarks_in_pid": "",
            "boccard_item_number": "",
            "tag_number": "",
            "designation": "",
            "category": "Uncategorized",
            "extra_data": {},
        }

        for index, value in enumerate(row):
            if index >= len(headers):
                continue
            _map_value(record, headers[index], value)

        desig = str(record.get("designation", "")).lower()
        if "valve" in desig or "actuator" in desig:
            record["category"] = "Valves & Actuators"
        elif "pump" in desig:
            record["category"] = "Pumps"
        elif "pipe" in desig or "fitting" in desig or "flange" in desig:
            record["category"] = "Piping & Fittings"
        elif "tank" in desig or "vessel" in desig:
            record["category"] = "Tanks & Vessels"
        elif "sensor" in desig or "transmitter" in desig or "gauge" in desig or "meter" in desig or "indicator" in desig:
            record["category"] = "Instruments"
        elif "motor" in desig or "electrical" in desig:
            record["category"] = "Electrical"
        elif desig:
            record["category"] = "Other"

        parsed_rows.append(record)

    return parsed_rows


def export_workbook_from_template(source_path, records):
    """
    Rebuilds an export by editing the ORIGINAL uploaded workbook in place
    (same sheet, same header row found during import) instead of creating a
    blank workbook. This keeps every other sheet, colors/fills, fonts,
    column widths, row heights, and merged cells exactly as they were in the
    source file - only the data cell values are refreshed.

    Returns a BytesIO with the exported .xlsx, or None if the original file
    can no longer be found on disk (caller should fall back to a plain export).
    """
    if not source_path or not os.path.exists(source_path):
        return None

    # Re-run the same sheet/header detection used at import time so we edit
    # the exact same place the data was read from.
    probe_workbook = load_workbook(source_path, data_only=True, read_only=True)
    _sheet, rows, header_index = _find_best_sheet(probe_workbook)
    if not rows:
        probe_workbook.close()
        return None
    sheet_title = _sheet.title
    headers = [str(header).strip() if header is not None else "" for header in rows[header_index]]
    probe_workbook.close()

    field_map = [(header, _resolve_field_for_header(header)) for header in headers]

    # Reload normally (styles/formulas/other sheets intact, fully editable).
    workbook = load_workbook(source_path)
    worksheet = workbook[sheet_title]

    header_row_number = header_index + 1
    original_last_row = worksheet.max_row
    template_style_row = header_row_number + 1 if original_last_row > header_row_number else None

    def apply_template_style(target_cell):
        if template_style_row is None:
            return
        template_cell = worksheet.cell(row=template_style_row, column=target_cell.column)
        target_cell._style = copy(template_cell._style)

    for offset, record in enumerate(records):
        row_number = header_row_number + 1 + offset
        is_new_row = row_number > original_last_row
        extra_data = record.get("extra_data") or {}
        for col_index, (header, field) in enumerate(field_map, start=1):
            cell = worksheet.cell(row=row_number, column=col_index)
            if is_new_row:
                apply_template_style(cell)
            if field:
                cell.value = record.get(field) or None
            elif header:
                cell.value = extra_data.get(header) or None
            # Columns with a blank/unrecognized header are left untouched.

    # If the dataset shrank (rows deleted in-app), clear leftover old rows
    # instead of leaving stale data behind.
    for row_number in range(header_row_number + 1 + len(records), original_last_row + 1):
        for col_index in range(1, len(field_map) + 1):
            worksheet.cell(row=row_number, column=col_index).value = None

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
